# sheets/final_forecast.py
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

MONTH_NAMES = ["", "січень", "лютий", "березень", "квітень", "травень", "червень",
               "липень", "серпень", "вересень", "жовтень", "листопад", "грудень"]


def create_sheet_final_forecast(workbook, params):
    sheet_name = "Фінальний прогноз"
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)

    # === Параметри ===
    model_year         = params["model_year"]
    headers            = params["input_headers"]
    trend_forecasts    = params["trend_forecasts"]
    seasonal_coeffs    = params["seasonal_coeffs"]
    factors_data       = params.get("factors_data", [])
    range_start_col    = params["range_start_col"]

    # === Фактори ===
    header_normalized = {h.strip().lower().replace(" ", ""): h for h in headers}
    factors_by_header = {}
    for f in factors_data:
        key = f["header"].strip().lower().replace(" ", "")
        if key in header_normalized:
            original = header_normalized[key]
            factors_by_header.setdefault(original, []).append({
                "desc": f["description"],
                "type": f["type"],
                "values": f["data"]
            })

    # === Розміри блоків ===
    block_sizes_no_sep = []
    for header in headers:
        factors_count = len(factors_by_header.get(header, []))
        block_sizes_no_sep.append(2 + factors_count + 1)  # Тренд + Сезонність + Фактори + Фінальний

    total_cols = 5
    for i, h in enumerate(headers):
        total_cols += block_sizes_no_sep[i]
        if h != headers[-1]:
            total_cols += 1  # роздільник

    # === Динамічні номери рядків ===
    HEADER_MAIN_ROW       = 1    # "Фінальний прогноз на 2025 рік"
    EMPTY_ROW             = 2
    REGION_HEADER_ROW     = 3    # Заголовки діапазонів даних
    COLUMN_HEADER_ROW     = 4    # Тренд, З урахуванням сезонності, Фінальний прогноз
    FIRST_DATA_ROW        = 5    # перший місяць (січень)

    # === Головний заголовок ===
    ws.cell(HEADER_MAIN_ROW, 1, f"Фінальний прогноз на {model_year} рік")
    ws.merge_cells(start_row=HEADER_MAIN_ROW, start_column=1,
                   end_row=HEADER_MAIN_ROW, end_column=total_cols)
    ws[f"A{HEADER_MAIN_ROW}"].font = Font(bold=True, size=14)
    ws[f"A{HEADER_MAIN_ROW}"].alignment = Alignment(horizontal="center", vertical="center")

    ws.append([])  # порожній рядок

    # === Рядок 3 — назви діапазонів даних (регіонів) ===
    cur_col = 6
    for idx, header in enumerate(headers):
        size = block_sizes_no_sep[idx]
        start = cur_col
        end   = cur_col + size - 1
        ws.merge_cells(start_row=REGION_HEADER_ROW, start_column=start,
                       end_row=REGION_HEADER_ROW, end_column=end)
        cell = ws.cell(REGION_HEADER_ROW, start, header)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cur_col += size + (1 if idx < len(headers)-1 else 0)

    # === Рядок 4 — детальні заголовки ===
    header_row = ["Рік", "Місяць", "Назва місяця", "Номер місяця", ""]
    for header in headers:
        factors = factors_by_header.get(header, [])
        header_row += ["Тренд", "З урахуванням сезонності"]
        for f in factors:
            header_row.append(f"{f['desc']} ({f['type']})")
        header_row.append("Фінальний прогноз")
        if header != headers[-1]:
            header_row.append("")
    ws.append(header_row)

    # === 12 місяців прогнозу ===
    final_forecast_by_col = {}
    for month_num in range(1, 13):
        row_values = [model_year, month_num, MONTH_NAMES[month_num], month_num, ""]
        for idx, header in enumerate(headers):
            col_idx = range_start_col + idx
            trend = trend_forecasts.get(col_idx, [0]*12)[month_num-1] or 0.0
            coeff = seasonal_coeffs.get((month_num, col_idx), 1.0)
            seasonal = round(trend * coeff, 2)

            row_values += [trend, seasonal]

            final_val = seasonal
            for f in factors_by_header.get(header, []):
                val = f["values"][month_num-1]
                if val is not None:
                    final_val = round(final_val * val, 2) if f["type"] == "коефіцієнт" else round(final_val + val, 2)
                row_values += [val if val is not None else ""]

            row_values += [final_val]
            if idx < len(headers) - 1:
                row_values += [""]

            final_forecast_by_col.setdefault(col_idx, []).append(final_val)

        ws.append(row_values)

    # === Стилі ===
    bold   = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap   = Alignment(horizontal="center", vertical="center", wrap_text=True)

    orange    = PatternFill("solid", fgColor="FF8C00")
    dark_blue = PatternFill("solid", fgColor="1F4E79")
    gray      = PatternFill("solid", fgColor="D3D3D3")
    light     = PatternFill("solid", fgColor="F0F0F0")
    blue      = PatternFill("solid", fgColor="DDEBF7")

    # Рядок діапазонів даних (регіонів)
    for cell in ws[REGION_HEADER_ROW]:
        if cell.value in headers:
            cell.fill = dark_blue
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.alignment = wrap

    # Детальні заголовки
    for cell in ws[COLUMN_HEADER_ROW]:
        if cell.value:
            cell.font = bold
            cell.alignment = wrap
            if cell.column <= 5:
                cell.fill = orange
            elif "Тренд" in str(cell.value):
                cell.fill = gray
            elif "сезонност" in str(cell.value):
                cell.fill = light
            elif "Фінальний" in str(cell.value):
                cell.fill = blue

    ws.row_dimensions[REGION_HEADER_ROW].height   = 40
    ws.row_dimensions[COLUMN_HEADER_ROW].height   = 100

    # === АВТОШИРИНА ТІЛЬКИ ПО ДАНИМ (рядки з прогнозами) ===
    data_start_row = FIRST_DATA_ROW
    data_end_row   = FIRST_DATA_ROW + 11  # 12 місяців

    column_widths = {}
    for row in ws.iter_rows(min_row=data_start_row, max_row=data_end_row):
        for cell in row:
            if cell.value is not None:
                length = len(str(cell.value))
                col = cell.column_letter
                column_widths[col] = max(column_widths.get(col, 0), length)

    for col, length in column_widths.items():
        ws.column_dimensions[col].width = max(8, min(length + 2, 20))

    # Фіксовані колонки A–E
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 5

    # Рамки
    thin = Side(border_style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=REGION_HEADER_ROW, max_row=ws.max_row,
                            min_col=1, max_col=total_cols):
        for cell in row:
            cell.border = border

    # Формат чисел
    for row in ws.iter_rows(min_row=data_start_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            if cell.value is not None:
                cell.alignment = center

    return {"final_forecast_by_col": final_forecast_by_col}