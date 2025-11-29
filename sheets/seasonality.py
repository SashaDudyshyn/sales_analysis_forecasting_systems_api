# sheets/seasonality.py
from openpyxl.styles import Font, Alignment, PatternFill
from collections import defaultdict
import numpy as np

MONTH_NAMES = [
    "", "січень", "лютий", "березень", "квітень", "травень", "червень",
    "липень", "серпень", "вересень", "жовтень", "листопад", "грудень"
]


def create_sheet_seasonality(workbook, params, smoothed_data):
    sheet_name = "Виключення сезонності"
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)

    # === Параметри ===
    col_start = params["range_start_col"]
    col_end = params["range_end_col"]
    input_headers = params.get("input_headers", [])
    years = params["years"]
    months = params["months"]
    total_months = len(years)
    data_cols = len(input_headers)

    # === 1. Розрахунок сезонних коефіцієнтів ===
    month_sums = defaultdict(lambda: {c: 0.0 for c in smoothed_data})
    month_counts = defaultdict(lambda: {c: 0 for c in smoothed_data})

    for i, m in enumerate(months):
        for c in smoothed_data:
            val = smoothed_data[c][i]
            if val is not None:
                month_sums[m][c] += val
                month_counts[m][c] += 1

    # Середнє по місяцях
    month_avg = {}
    for m in range(1, 13):
        month_avg[m] = {}
        for c in smoothed_data:
            if month_counts[m][c] > 0:
                month_avg[m][c] = month_sums[m][c] / month_counts[m][c]
            else:
                month_avg[m][c] = None

    # Загальне середнє по колонці
    overall_avg = {}
    for c in smoothed_data:
        valid_vals = [v for v in smoothed_data[c] if v is not None]
        overall_avg[c] = np.mean(valid_vals) if valid_vals else None

    # Ненормовані коефіцієнти
    unnormalized = {}
    for m in range(1, 13):
        unnormalized[m] = {}
        for c in smoothed_data:
            if month_avg[m][c] is not None and overall_avg[c] and overall_avg[c] != 0:
                unnormalized[m][c] = month_avg[m][c] / overall_avg[c]
            else:
                unnormalized[m][c] = 1.0

    # Нормалізація: сума за рік = 12
    normalized = {}
    for c in smoothed_data:
        S = sum(unnormalized[m][c] for m in range(1, 13))
        N = 12.0 / S if S != 0 else 1.0
        for m in range(1, 13):
            normalized[(m, c)] = round(unnormalized[m][c] * N, 4)

    # === Десезоналізація (основний результат!) ===
    deseasoned_data = {}  # {col_index: [значення по періодах]}
    deseasoned_by_row = {}  # для запису на аркуш

    for i in range(total_months):
        m = months[i]
        deseasoned_by_row[i] = {}
        for c in smoothed_data:
            coeff = normalized.get((m, c), 1.0)
            val = smoothed_data[c][i]
            deseasoned_val = val / coeff if coeff != 0 and val is not None else None
            deseasoned_by_row[i][c] = round(deseasoned_val, 2) if deseasoned_val is not None else None

            if c not in deseasoned_data:
                deseasoned_data[c] = []
            deseasoned_data[c].append(deseasoned_by_row[i][c])

    # === Позиції колонок ===
    smoothed_start = 5
    unnorm_month_start = smoothed_start + data_cols + 2
    unnorm_coeff_start = unnorm_month_start + 1
    norm_month_start = unnorm_coeff_start + data_cols + 2
    norm_coeff_start = norm_month_start + 1
    deseasoned_start = norm_coeff_start + data_cols + 2

    # === Запис заголовків ===
    def add_title(start_col, end_col, text):
        cell = ws.cell(1, start_col, text)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    add_title(1, smoothed_start + data_cols - 1, "Згладжені дані")
    add_title(unnorm_month_start, unnorm_coeff_start + data_cols - 1, "Ненормовані сезонні коефіцієнти")
    add_title(norm_month_start, norm_coeff_start + data_cols - 1, "Нормовані сезонні коефіцієнти")
    add_title(deseasoned_start, deseasoned_start + 3 + data_cols, "Десезоналізовані дані")

    ws.append([])  # рядок 2

    # Рядок 3 — детальні заголовки
    header_row = (
        ["Рік", "Місяць", "Назва місяця", "Номер"] + input_headers +
        ["", ""] +
        ["Місяць"] + input_headers +
        ["", ""] +
        ["Місяць"] + input_headers +
        ["", ""] +
        ["Рік", "Місяць", "Назва місяця", "Номер"] + input_headers
    )
    ws.append(header_row)

    # === Заповнення ===
    for i in range(total_months):
        row = 4 + i
        m = months[i]
        ws.cell(row, 1, years[i])
        ws.cell(row, 2, m)
        ws.cell(row, 3, MONTH_NAMES[m])
        ws.cell(row, 4, i + 1)

        # Згладжені
        for idx, c in enumerate(range(col_start, col_end + 1)):
            val = smoothed_data[c][i]
            ws.cell(row, smoothed_start + idx, round(val, 2) if val else None)

        # Десезоналізовані
        ws.cell(row, deseasoned_start, years[i])
        ws.cell(row, deseasoned_start + 1, m)
        ws.cell(row, deseasoned_start + 2, MONTH_NAMES[m])
        ws.cell(row, deseasoned_start + 3, i + 1)
        for idx, c in enumerate(range(col_start, col_end + 1)):
            val = deseasoned_by_row[i].get(c)
            ws.cell(row, deseasoned_start + 4 + idx, val)

        # Коефіцієнти (перші 12 місяців)
        if i < 12:
            mm = i + 1
            ws.cell(row, unnorm_month_start, MONTH_NAMES[mm])
            ws.cell(row, norm_month_start, MONTH_NAMES[mm])
            for idx, c in enumerate(range(col_start, col_end + 1)):
                ws.cell(row, unnorm_coeff_start + idx, round(unnormalized[mm][c], 4))
                ws.cell(row, norm_coeff_start + idx, normalized.get((mm, c), 1.0))

    # === Стилі ===
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    orange = PatternFill("solid", fgColor="FF8C00")
    gray = PatternFill("solid", fgColor="D3D3D3")

    for cell in ws[3]:
        cell.font = bold
        cell.alignment = center
        if cell.column <= 4 or cell.column >= deseasoned_start:
            cell.fill = orange
        else:
            cell.fill = gray

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.alignment = center

    # === АВТОШИРИНА ===
    for col_letter in ws.column_dimensions:
        max_length = 10
        for cell in ws[col_letter]:
            if cell.value and cell.coordinate not in ws.merged_cells:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # === ПОВЕРТАЄМО ДЕСЕЗОНАЛІЗОВАНІ ДАНІ ДЛЯ ПРОГНОЗУ ===
    return {
        "deseasoned_data": deseasoned_data,
        "seasonal_coeffs": normalized,
    }