# sheets/visualization.py
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill


def create_combined_visualization_from_columns(
    workbook,
    years,
    months,
    raw_data_dict,
    smoothed_dict,
    deseasoned_dict,
    forecast_dict,
    column_headers,
    model_year
):
    sheet_name = "Візуалізація"
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)

    current_row = 1

    # Отримуємо правильний порядок колонок: ключі з raw_data_dict (7,8,9,10)
    data_columns = sorted(raw_data_dict.keys())  # [7, 8, 9, 10]

    for col_idx in data_columns:
        # Знаходимо індекс у списку заголовків (0,1,2,3)
        header_index = data_columns.index(col_idx)
        header_name = column_headers[header_index]

        raw = raw_data_dict[col_idx]
        smooth = smoothed_dict.get(col_idx, [])
        deseas = deseasoned_dict.get(col_idx, [])
        final_fc = forecast_dict.get(col_idx, [])

        n_hist = len(years)

        #Заголовок
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        title_cell = ws.cell(current_row, 1, f"Динаміка та прогноз: {header_name}")
        title_cell.font = Font(bold=True, size=16, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[current_row].height = 45
        current_row += 1

        # Заголовки таблиці
        headers = ["Період", "Сирі дані", "Згладжені", "Тренд", "Фінальний прогноз"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(current_row, c, h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        current_row += 1

        # Дані
        data_start_row = current_row
        for i in range(n_hist):
            period = f"{int(years[i])}-{int(months[i]):02d}"
            ws.append([
                period,
                raw[i] if i < len(raw) else None,
                smooth[i] if i < len(smooth) else None,
                deseas[i] if i < len(deseas) else None,
                None
            ])

        for m in range(1, 13):
            fc_val = final_fc[m-1] if m-1 < len(final_fc) else None
            ws.append([f"{model_year}-{m:02d}", None, None, None, fc_val])

        data_end_row = ws.max_row

        # Автоширина
        for col in range(1, 6):
            column = get_column_letter(col)
            max_len = 0
            for r in range(data_start_row, data_end_row + 1):
                val = ws.cell(r, col).value
                if val is not None:
                    val_str = f"{val:,.0f}".replace(",", " ") if isinstance(val, (int, float)) else str(val)
                    max_len = max(max_len, len(val_str))
            width = max(min(max_len + 2, 25), 10)
            ws.column_dimensions[column].width = width

        # Графік
        chart = LineChart()
        chart.title = f"Прогноз: {header_name}"
        chart.style = 27
        chart.height = 18
        chart.width = 34
        chart.x_axis.title = "Період"
        chart.y_axis.title = "Обсяг"
        chart.legend.position = "b"

        data_ref = Reference(ws, min_col=2, max_col=5, min_row=current_row-1, max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        colors = ["1F4E79", "ED7D31", "A5A5A5", "70AD47"]
        for i, s in enumerate(chart.series):
            s.graphicalProperties.line.solidFill = colors[i]
            s.graphicalProperties.line.width = 28000
            if i == 3:
                s.graphicalProperties.line.dashStyle = "dash"
                s.graphicalProperties.line.width = 35000
            s.marker.symbol = "circle"
            s.marker.size = 7

        ws.add_chart(chart, f"G{data_start_row}")

        #Роздільник
        current_row = data_end_row + 5
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
        sep = ws.cell(current_row, 1)
        sep.fill = PatternFill("solid", fgColor="1F4E79")
        ws.row_dimensions[current_row].height = 4
        current_row += 2

    return ws
