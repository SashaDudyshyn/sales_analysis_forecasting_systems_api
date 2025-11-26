# sheets/smoothed_data.py
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

MONTH_NAMES = [
    "", "січень", "лютий", "березень", "квітень", "травень", "червень",
    "липень", "серпень", "вересень", "жовтень", "листопад", "грудень"
]


def create_sheet_smoothed_data(workbook, params):
    sheet_name = "Згладжені дані"
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)

    active_sheet = params["active_sheet"]
    col_start = params["range_start_col"]
    col_end = params["range_end_col"]
    row_first = params["row_first_data"]
    row_last = params["row_last_data"]
    k = params.get("k", 2)  # ← тепер реально використовується!

    input_headers = params["input_headers"]
    data_cols = len(input_headers)

    # === Читання сирих даних ===
    raw_data = {c: [] for c in range(col_start, col_end + 1)}
    years = []
    months = []

    for row in active_sheet.iter_rows(min_row=row_first, max_row=row_last):
        year = row[1].value      # колонка B
        month = row[3].value     # колонка D
        if year is None or month is None:
            continue
        years.append(year)
        months.append(month)

        for c in range(col_start, col_end + 1):
            val = row[c - 1].value
            raw_data[c].append(float(val) if val is not None else None)

    # === ЗГЛАДЖУВАННЯ: центроване ковзне середнє з вікном 2k+1 ===
    n = len(years)
    smoothed = {}

    for c in raw_data:
        values = raw_data[c]
        smoothed[c] = []

        for i in range(n):
            if values[i] is None:
                smoothed[c].append(None)
                continue

            # Визначаємо межі вікна
            if i < k:
                start = 0
                end = min(2 * i + 1, n)
            elif i >= n - k:
                start = max(0, 2 * i - n + 1)
                end = n
            else:
                start = i - k
                end = i + k + 1

            window = [v for v in values[start:end] if v is not None]
            avg = sum(window) / len(window) if window else None
            smoothed[c].append(round(avg, 2) if avg is not None else None)

    # === РОЗМІТКА АРКУША ===
    block_width = 4 + data_cols  # Рік, Місяць, Назва, Номер + дані

    # Лівий блок: ВХІДНІ ДАНІ
    left_title = ws.cell(1, 1, "ВХІДНІ ДАНІ")
    ws.merge_cells(
        start_row=1, start_column=1,
        end_row=1, end_column=block_width
    )

    # Правий блок: ЗГЛАДЖЕНІ ДАНІ
    right_start_col = block_width + 3  # +2 відступи + 1
    right_end_col = right_start_col + block_width - 1

    right_title = ws.cell(1, right_start_col, f"ЗГЛАДЖЕНІ ДАНІ (k={k})")
    ws.merge_cells(
        start_row=1, start_column=right_start_col,
        end_row=1, end_column=right_end_col
    )

    # Порожній рядок
    ws.append([])

    # Заголовки рядка 3
    header_row = (
        ["Рік", "Місяць", "Назва місяця", "Номер місяця"] + input_headers +
        ["", ""] +
        ["Рік", "Місяць", "Назва місяця", "Номер місяця"] + input_headers
    )
    ws.append(header_row)

    # Дані
    for i in range(n):
        month_name = MONTH_NAMES[months[i]]
        row = [
            years[i], months[i], month_name, i + 1,
        ] + [raw_data[c][i] for c in range(col_start, col_end + 1)] + [
            "", "",
            years[i], months[i], month_name, i + 1
        ] + [smoothed[c][i] for c in range(col_start, col_end + 1)]
        ws.append(row)

    # === СТИЛІ ===
    bold_large = Font(bold=True, size=14)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    fill_orange = PatternFill("solid", fgColor="FF8C00")
    fill_gray = PatternFill("solid", fgColor="D3D3D3")

    # Великі заголовки
    left_title.font = bold_large
    left_title.alignment = center
    right_title.font = bold_large
    right_title.alignment = center

    # Заголовки рядка 3
    for cell in ws[3]:
        cell.font = bold
        cell.alignment = center
        col = cell.column
        if col <= 4 or col >= right_start_col:  # мета-колонки обох блоків
            cell.fill = fill_orange
        else:
            cell.fill = fill_gray

    # Центрування всіх чисел
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                cell.alignment = center

    # === АВТОШИРИНА ===
    from openpyxl.utils import get_column_letter

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_idx)

        max_length = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            if getattr(cell, "is_merged", False):
                continue
            max_length = max(max_length, len(str(cell.value)))

        # Мінімальна ширина, щоб не було "порожніх" колонок
        ws.column_dimensions[col_letter].width = min(max_length + 2 if max_length > 0 else 12, 50)

    # Повертаємо дані для наступних кроків
    return {
        "smoothed_data": smoothed,
        "years": years,
        "months": months,
    }