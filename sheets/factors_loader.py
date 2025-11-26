# sheets/factors_loader.py
from openpyxl.utils import column_index_from_string, get_column_letter

def load_factors_data(workbook, params):
    """
    Читає дані з аркуша "Фактори впливу"
    Повертає список словників:
    [
        {
            "description": "Температура",
            "type": "коефіцієнт",  # або "одиниці"
            "header": "Темп. пов.",
            "year_col": 2023,
            "month_col": 1,
            "data": [0.95, 1.02, ...]  # по місяцях
        },
        ...
    ]
    """
    ws_factor = workbook[params["sheet_factor"]]

    year_col_letter = params["factor_column_year"]
    month_col_letter = params["factor_column_month"]
    range_str = params["factor_row_range_data"]
    desc_row = params["factor_row_description"]
    type_row = params["factor_row_type"]
    title_row = params["factor_row_title"]
    first_data_row = params["factor_row_first_data"]
    last_data_row = params["factor_row_last_data"]

    start_col = column_index_from_string(range_str.split("-")[0])
    end_col = column_index_from_string(range_str.split("-")[1])

    factors = []

    for col in range(start_col, end_col + 1):
        description = ws_factor.cell(desc_row, col).value or ""
        factor_type = ws_factor.cell(type_row, col).value
        header = ws_factor.cell(title_row, col).value or f"Фактор {get_column_letter(col)}"

        if not factor_type or factor_type not in ["коефіцієнт", "одиниці"]:
            continue  # або можна кидати помилку

        data = []
        for row in range(first_data_row, last_data_row + 1):
            val = ws_factor.cell(row, col).value
            data.append(float(val) if val is not None else None)

        factors.append({
            "description": str(description).strip(),
            "type": factor_type.lower(),
            "header": str(header),
            "data": data,
        })

    return factors