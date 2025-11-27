# sheets/start_parameters.py
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_sheet_start_parameters(workbook, params):
    sheet_name = "Початкові налаштування"
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])

    ws = workbook.create_sheet(title=sheet_name)

    # Формуємо список регіонів
    headers = params.get("input_headers", [])
    headers_str = ", ".join(headers) if headers else "—"

    # === СПИСОК РЯДКІВ ===
    rows = [
        ["Параметр", "Значення"],
        ["Файл", params["filename"]],
        ["Аркуш зі статистикою", params["sheet_stat"]],
        ["Аркуш з факторами впливу", params["sheet_factor"]],
        ["Рік прогнозу", params["model_year"]],
        ["", ""],
        ["Налаштування статистичних даних", ""],
        ["Колонка року", params["column_year"]],
        ["Колонка місяця", params["column_month"]],
        ["Діапазон даних", params["range_data"]],
        ["Рядок заголовків", params["row_title"]],
        ["Перший рядок даних", params["row_first_data"]],
        ["Останній рядок даних", params["row_last_data"]],
        ["Коефіцієнт згладжування (k)", params["k"]],
        ["Набори даних", headers_str],
        ["", ""],
        ["Налаштування факторів впливу", ""],
        ["Колонка року (фактори)", params["factor_column_year"]],
        ["Колонка місяця (фактори)", params["factor_column_month"]],
        ["Діапазон значень факторів", params["factor_row_range_data"]],
        ["Рядок опису", params["factor_row_description"]],
        ["Рядок типу", params["factor_row_type"]],
        ["Рядок назв факторів", params["factor_row_title"]],
        ["Перший рядок даних (фактори)", params["factor_row_first_data"]],
        ["Останній рядок даних (фактори)", params["factor_row_last_data"]],
    ]

    # === ЗАПИСУЄМО ВСІ РЯДКИ СПЕРШУ ===
    for r in rows:
        ws.append(r)

    # === СТИЛІ ТА ФОРМАТУВАННЯ ===
    bold_white = Font(bold=True, color="FFFFFF", size=13)
    bold_blue  = Font(bold=True, color="1F4E79", size=12)
    regular    = Font(size=11)
    value_font = Font(size=11, bold=True, color="1F4E79")

    header_fill  = PatternFill("solid", fgColor="1F4E79")
    section_fill = PatternFill("solid", fgColor="DDEBF7")
    border = Border(left=Side("thin"), right=Side("thin"),
                    top=Side("thin"), bottom=Side("thin"))

    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_indent = Alignment(horizontal="left", vertical="center", indent=1)

    # Заголовок таблиці
    ws["A1"].font = bold_white
    ws["B1"].font = bold_white
    ws["A1"].fill = header_fill
    ws["B1"].fill = header_fill
    ws["A1"].alignment = center_wrap
    ws["B1"].alignment = center_wrap

    # === ОБ'ЄДНАННЯ КЛІТИНОК — РОБИМО ПІСЛЯ ЗАПИСУ ВСІХ ДАНИХ ===
    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row_idx, 1)
        if cell_a.value and "Налаштування" in str(cell_a.value):
            # Спочатку очищаємо B-клітинку (щоб не було конфлікту)
            ws.cell(row_idx, 2).value = None
            # Тепер безпечно об'єднуємо
            ws.merge_cells(start_row=row_idx, start_column=1,
                           end_row=row_idx, end_column=2)
            cell_a.font = bold_blue
            cell_a.fill = section_fill
            cell_a.alignment = left_indent

    # Основні рядки (окрім об'єднаних)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        a_cell = row[0]
        b_cell = row[1]

        a_cell.font = regular
        b_cell.font = value_font
        a_cell.alignment = left_indent
        b_cell.alignment = center_wrap

        a_cell.border = border
        b_cell.border = border

    # Висота рядків
    ws.row_dimensions[1].height = 28
    for i in range(2, ws.max_row + 1):
        if "Налаштування" in str(ws.cell(i, 1).value or ""):
            ws.row_dimensions[i].height = 26
        elif ws.cell(i, 1).value == "Набори даних" and len(headers_str) > 80:
            ws.row_dimensions[i].height = 38
        else:
            ws.row_dimensions[i].height = 22

    # === АВТОШИРИНА ===
    for col_letter in ("A", "B"):
        max_len = 0
        column = ws[col_letter]
        for cell in column:
            if cell.value:
                length = len(str(cell.value))
                max_len = max(max_len, length)
        width = min(max_len + 4, 70)
        ws.column_dimensions[col_letter].width = width

    # Заморозка
    ws.freeze_panes = "A2"

    return ws