# sheets/forecast.py
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np

MONTH_NAMES = ["", "січень", "лютий", "березень", "квітень", "травень", "червень",
               "липень", "серпень", "вересень", "жовтень", "листопад", "грудень"]


def create_sheet_forecast(workbook, params, deseasoned_data: dict):
    sheet_name = "Прогноз"
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    ws = workbook.create_sheet(title=sheet_name)

    headers = params["input_headers"]
    years = params["years"]
    months = params["months"]
    model_year = params["model_year"]

    n_hist = len(years)
    n_forecast = 12
    total_periods = n_hist + n_forecast

    # — Лінійна регресія —
    x_hist = np.arange(1, n_hist + 1)
    x_forecast = np.arange(n_hist + 1, total_periods + 1)

    trends = {}
    for col_idx, values in deseasoned_data.items():
        y = np.array([v for v in values if v is not None], dtype=float)
        if len(y) < 2:
            A, B = 0.0, 0.0
        else:
            B, A = np.polyfit(np.arange(1, len(y) + 1), y, 1)
        trend_hist = (A + B * x_hist).round(2).tolist()
        forecast   = (A + B * x_forecast).round(2).tolist()

        trends[col_idx] = {
            "A": round(A, 2),
            "B": round(B, 2),
            "trend_hist": trend_hist,
            "forecast":   forecast,
        }

    # — Головний заголовок —
    title = "Модель лінійного тренду для згладжених даних з виключеною сезонною компонентою"
    total_data_cols = len(headers) * 2 + max(0, len(headers) - 1)   # 2 колонки на регіон + порожній між ними (крім останнього)
    total_cols = 5 + total_data_cols                               # 4 мета + 1 порожній після "Номер періоду" + дані
    ws.cell(1, 1, title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.append([])  # рядок 2

    # — Коефіцієнти (рядок 3) —
    coeff_start_col = 6
    current_col = coeff_start_col
    for i, header in enumerate(headers):
        col_idx = params["range_start_col"] + i
        t = trends.get(col_idx, {"A": "—", "B": "—"})
        txt = f"Коефіцієнти: intercept = {t['A']}, slope = {t['B']}"
        ws.cell(3, current_col, txt)
        ws.merge_cells(start_row=3, start_column=current_col, end_row=3, end_column=current_col + 1)
        cell = ws.cell(3, current_col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1f4e79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        current_col += 3  # 2 колонки даних + 1 порожній (для всіх, крім останнього — виправиться нижче)

    # — Забезпечуємо читабельність рядка з коефіцієнтами —
    for cell in ws[3]:
        if cell.value:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 45  # Excel сам підлаштує вище при відкритті

    # — Заголовки колонок (рядок 4) —
    header_row = ["Рік", "Місяць", "Назва місяця", "Номер періоду", ""]
    for i, h in enumerate(headers):
        header_row += [h, "ТРЕНД"]
        if i < len(headers) - 1:        # додаємо порожній стовпець між регіонами
            header_row += [""]
    ws.append(header_row)

    # Стилі заголовків
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    orange = PatternFill("solid", fgColor="FF8C00")
    gray = PatternFill("solid", fgColor="D3D3D3")
    light = PatternFill("solid", fgColor="F0F0F0")

    for cell in ws[4]:
        cell.font = bold
        cell.alignment = center
        col = cell.column
        if col <= 5:
            cell.fill = orange
        elif (col - 5) % 3 == 1:        # десезоналізовані
            cell.fill = light
        elif (col - 5) % 3 == 2:        # ТРЕНД
            cell.fill = gray

    # — Дані —
    for period in range(1, total_periods + 1):
        if period <= n_hist:
            i = period - 1
            year, month = years[i], months[i]
            month_name = MONTH_NAMES[month]
            is_forecast = False
        else:
            i = period - n_hist - 1
            year = model_year
            month = (i % 12) + 1
            month_name = MONTH_NAMES[month]
            is_forecast = True

        row = [year, month, month_name, period, ""]

        for idx, header in enumerate(headers):
            col_idx = params["range_start_col"] + idx
            t = trends.get(col_idx, {})

            if is_forecast:
                deseason_val = None
                trend_val = t.get("forecast", [None]*12)[i]
            else:
                deseason_val = deseasoned_data[col_idx][period - 1]
                trend_val = t.get("trend_hist", [None]*n_hist)[period - 1]

            row += [deseason_val, trend_val]
            if idx < len(headers) - 1:      # порожній стовпець між регіонами
                row += [""]

        ws.append(row)

    # — Форматування чисел та прогнозу —
    for row_cells in ws.iter_rows(min_row=5):
        for cell in row_cells:
            if isinstance(cell.value, (int, float)):
                cell.alignment = center
            if cell.row > 4 + n_hist:
                cell.font = Font(color="000080", bold=True)

    # — Автоширина (безпечна) —
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for cell in col_cells:
            if cell.value and cell.coordinate not in ws.merged_cells:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Рамки (опціонально)
    thin = Side(border_style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border