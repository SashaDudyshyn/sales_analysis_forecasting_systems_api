# main.py
from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from io import BytesIO

from models.excel_params import ExcelProcessParams
from sheets.start_parameters import create_sheet_start_parameters
from sheets.smoothed_data import create_sheet_smoothed_data
from sheets.seasonality import create_sheet_seasonality
from sheets.forecast import create_sheet_forecast
from sheets.factors_loader import load_factors_data
from sheets.final_forecast import create_sheet_final_forecast
from sheets.visualization import create_sheet_visualization

app = FastAPI(title="Прогноз продажів")

@app.post("/process-excel/")
async def process_excel(
    file: UploadFile = File(...),

    # === Статистичні дані ===
    column_year: str = Form("B"),
    column_month: str = Form("D"),
    range_data: str = Form("G-J"),
    row_title: int = Form(3),
    row_first_data: int = Form(4),
    row_last_data: int = Form(38),
    k: int = Form(2),

    # === Аркуші ===
    sheet_stat: str = Form("Статистичні дані"),
    sheet_factor: str = Form("Фактори впливу"),

    # === Фактори впливу ===
    factor_column_year: str = Form("B"),
    factor_column_month: str = Form("C"),
    factor_row_range_data: str = Form("E-F"),
    factor_row_description: int = Form(3),
    factor_row_type: int = Form(4),
    factor_row_title: int = Form(5),
    factor_row_first_data: int = Form(6),
    factor_row_last_data: int = Form(17),
):
    # Перевірка формату файлу
    if not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(400, "Підтримуються тільки файли .xlsx")

    content = await file.read()
    workbook = load_workbook(filename=BytesIO(content))

    # === ПОВНА ВАЛІДАЦІЯ ВСІХ ПАРАМЕТРІВ ===
    try:
        params = ExcelProcessParams(
            column_year=column_year,
            column_month=column_month,
            range_data=range_data,
            row_title=row_title,
            row_first_data=row_first_data,
            row_last_data=row_last_data,
            k=k,

            sheet_stat=sheet_stat,
            sheet_factor=sheet_factor,

            factor_column_year=factor_column_year,
            factor_column_month=factor_column_month,
            factor_row_range_data=factor_row_range_data,
            factor_row_description=factor_row_description,
            factor_row_type=factor_row_type,
            factor_row_title=factor_row_title,
            factor_row_first_data=factor_row_first_data,
            factor_row_last_data=factor_row_last_data,
        )
    except ValueError as e:
        raise HTTPException(422, f"Помилка валідації: {e}")

    # Перетворюємо в dict (зручніше передавати)
    params_dict = params.model_dump()

    # === Розрахунок колонок ===
    col_start = column_index_from_string(params.range_data.split("-")[0])
    col_end = column_index_from_string(params.range_data.split("-")[1])

    # === Читаємо правильні заголовки з аркуша sheet_stat ===
    try:
        stat_sheet = workbook[params_dict["sheet_stat"]]
    except KeyError:
        raise HTTPException(400, f"Аркуш '{params_dict['sheet_stat']}' не знайдено у файлі")

    correct_headers = []
    for c in range(col_start, col_end + 1):
        val = stat_sheet.cell(row=params.row_title, column=c).value
        header = str(val).strip() if val else f"Колонка {get_column_letter(c)}"
        correct_headers.append(header)

    # === params_dict ===
    params_dict.update({
        "workbook": workbook,
        "active_sheet": stat_sheet,           # ← КРИТИЧНО ВАЖЛИВО!
        "range_start_col": col_start,
        "range_end_col": col_end,
        "input_headers": correct_headers,
        "filename": file.filename,
    })

    # === розрахунок року прогнозу ===
    # Беремо останній рік зі статистичних даних і додаємо +1
    last_year = None
    year_col_idx = column_index_from_string(params.column_year)

    for row in stat_sheet.iter_rows(
            min_row=params.row_first_data,
            max_row=params.row_last_data,
            min_col=year_col_idx,
            max_col=year_col_idx,
            values_only=True
    ):
        val = row[0]
        if val is not None:
            try:
                last_year = int(val)
            except (ValueError, TypeError):
                continue

    if last_year is None:
        raise HTTPException(400, "Не знайдено жодного року у колонці з роками")

    model_year = last_year + 1
    params_dict["model_year"] = model_year

    # === 1. Аркуш з параметрами ===
    create_sheet_start_parameters(workbook, params_dict)

    # === 2. Згладжені дані ===
    smoothed_result = create_sheet_smoothed_data(workbook, params_dict)

    # === 3. Підготовка до сезонності ===
    final_params = {
        **params_dict,
        "years": smoothed_result["years"],
        "months": smoothed_result["months"],
    }

    # === 4. Виключення сезонності ===
    create_sheet_seasonality(workbook, final_params, smoothed_result["smoothed_data"])
    seasonality_result = create_sheet_seasonality(workbook, final_params, smoothed_result["smoothed_data"])
    final_params.update({
        "deseasoned_data": seasonality_result["deseasoned_data"],
        "seasonal_coeffs": seasonality_result["seasonal_coeffs"],
    })


    # === 5. Тренд ===
    create_sheet_forecast(workbook, final_params, seasonality_result["deseasoned_data"])

    # === 6. Завантаження факторів впливу ===
    try:
        factors_data = load_factors_data(workbook, params_dict)
        final_params["factors_data"] = factors_data
    except Exception as e:
        raise HTTPException(500, f"Помилка читання факторів впливу: {e}")

    # === 7. Фінальний прогноз ===
    forecast_result = create_sheet_forecast(workbook, final_params, seasonality_result["deseasoned_data"])
    final_params["trend_forecasts"] = forecast_result["trend_forecasts"]

    final_result = create_sheet_final_forecast(workbook, final_params)
    final_forecast_by_col = final_result["final_forecast_by_col"]

    # === ЗБИРАЄМО ВСІ ДАНІ ДЛЯ ВІЗУАЛІЗАЦІЇ (все зі словників, НЕ з Pydantic!) ===

    # 1. Роки та місяці
    years_list = []
    months_list = []

    year_col = column_index_from_string(params_dict["column_year"])
    month_col = column_index_from_string(params_dict["column_month"])

    for row in stat_sheet.iter_rows(
            min_row=params_dict["row_first_data"],
            max_row=params_dict["row_last_data"],
            min_col=year_col,
            max_col=month_col,
            values_only=True
    ):
        if len(row) >= 2 and row[0] is not None and row[1] is not None:
            try:
                years_list.append(int(row[0]))
                months_list.append(int(row[1]))
            except (ValueError, TypeError):
                continue

    # 2. Сирі дані
    raw_data_dict = {}
    start_col = column_index_from_string(params_dict["range_data"].split("-")[0])
    end_col = column_index_from_string(params_dict["range_data"].split("-")[1])

    for col_idx in range(start_col, end_col + 1):
        col_values = []
        for cell in stat_sheet.iter_rows(
                min_row=params_dict["row_first_data"],
                max_row=params_dict["row_last_data"],
                min_col=col_idx,
                max_col=col_idx,
                values_only=True
        ):
            val = cell[0]
            try:
                col_values.append(float(val)) if val is not None else col_values.append(None)
            except (ValueError, TypeError):
                col_values.append(None)
        raw_data_dict[col_idx] = col_values

    # 3. Назви регіонів (input_headers) — беремо з params_dict!
    input_headers = params_dict.get("input_headers", [])

    # === ПАРАМЕТРИ ДЛЯ ВІЗУАЛІЗАЦІЇ — ТІЛЬКИ СЛОВНИК! ===
    viz_params = {
        "model_year": model_year,
        "input_headers": input_headers,  # ← з params_dict!
        "range_start_col": start_col,

        "years": years_list,
        "months": months_list,

        "raw_data": raw_data_dict,
        "smoothed_data": smoothed_result["smoothed_data"],
        "deseasoned_data": seasonality_result["deseasoned_data"],
        "trend_forecasts": forecast_result["trend_forecasts"],
        "final_forecast": final_forecast_by_col,
    }

    # === СТВОРЮЄМО ВКЛАДКУ З ГРАФІКОМ ===
    create_sheet_visualization(workbook, viz_params)


    # === Повертаємо готовий файл ===
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=processed_{file.filename}"}
    )